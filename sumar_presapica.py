"""
Radar DG - sumar Base Presapica.xlsx al catalogo
====================================================
Este Excel tiene 3 hojas que se suman al catalogo (ademas de los ~1,324
SKUs vigentes), para que el radar compare tambien contra:

  - "China SS27"  -> candidatos de importacion de China (CON fotos)
  - "Presapica"   -> lo ya aprobado/comprado antes de viajar a SAPICA (CON fotos)
  - "PRESS27"     -> pedidos de importacion en proceso (SIN fotos: solo
                     datos de proveedor/costo, no compite visualmente
                     todavia, pero queda como referencia)

No necesita internet: las fotos de "China SS27" y "Presapica" estan
incrustadas adentro del mismo Excel.

Uso:
    python sumar_presapica.py "Base Presapica.xlsx" catalog_index.json
"""
import sys
import json
import base64
import io
import openpyxl
from PIL import Image
import imagehash

THUMB_SIZE = (220, 220)
JPEG_QUALITY = 70

# nombre de hoja -> (fuente, tiene_fotos)
HOJAS_CON_FOTO = {
    "China SS27": "china_ss27",
    "Presapica": "presapica",
}


def normaliza_categoria(cat) -> str:
    if not cat:
        return "Sin categoría"
    return str(cat).strip().title()


def imagen_a_thumb(data: bytes):
    im = Image.open(io.BytesIO(data)).convert("RGB")
    phash = str(imagehash.phash(im, hash_size=16))
    colorhash = str(imagehash.colorhash(im, binbits=3))
    thumb = im.copy()
    thumb.thumbnail(THUMB_SIZE)
    buf = io.BytesIO()
    thumb.save(buf, format="JPEG", quality=JPEG_QUALITY)
    thumb_b64 = base64.b64encode(buf.getvalue()).decode()
    return phash, colorhash, thumb_b64


def procesar_hoja_con_fotos(ws, fuente):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}

    img_por_fila = {}
    for im in getattr(ws, "_images", []):
        fila_excel = im.anchor._from.row + 1  # 1-indexed, para que coincida con iter_rows
        try:
            img_por_fila[fila_excel] = im._data()
        except Exception:
            pass

    registros = []
    for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue

        def g(col):
            i = idx.get(col)
            return row[i] if i is not None and i < len(row) else None

        sku = g("SKU")
        if not sku:
            continue

        rec = {
            "sku": str(sku).strip(),
            "categoria": normaliza_categoria(g("Categoria") or g("Categoría")),
            "subcategoria": g("Subcategoria") or g("Subcategoría"),
            "inventario": g("Total Inv ") or g("Total Inv") or 0,
            "ventas": g("VT") or 0,
            "sell_through_pct": round((g("ST") or 0) * 100) if isinstance(g("ST"), float) and g("ST") <= 1 else (g("ST") or 0),
            "costo": g("Costo"),
            "precio": g("PV"),
            "fuente": fuente,
        }

        data = img_por_fila.get(n)
        if data:
            try:
                phash, colorhash, thumb_b64 = imagen_a_thumb(data)
                rec["phash"] = phash
                rec["colorhash"] = colorhash
                rec["thumb_b64"] = thumb_b64
            except Exception as e:
                print(f"  [{fuente}] no pude procesar la foto de {sku}: {e}")
                rec["sin_imagen_procesada"] = True
        else:
            rec["sin_imagen_procesada"] = True

        registros.append(rec)
    return registros


def procesar_press27(ws):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}

    registros = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        def g(col):
            i = idx.get(col)
            return row[i] if i is not None and i < len(row) else None

        codigo_dg = g("CODIGO DG")
        codigo_color = g("CODIGO COLOR")
        oc = g("OC")
        if not (codigo_dg or codigo_color):
            continue

        sku_sintetico = f"PRESS27-{oc}-{codigo_color or codigo_dg}"

        registros.append({
            "sku": sku_sintetico,
            "categoria": normaliza_categoria(g("CATEGORIA")),
            "subcategoria": g("SUBCATEGORIA"),
            "proveedor": g("PROVEEDOR"),
            "color": g("COLOR "),
            "costo": g("COSTO NUEVO") or g("COSTO USD"),
            "precio": g("PV NUEVO"),
            "margen": g("MARGEN"),
            "inventario": g("PARES") or 0,
            "ventas": 0,
            "sell_through_pct": 0,
            "status": g("STATUS "),
            "comentarios": g("COMENTARIOS"),
            "fuente": "press27_sin_foto",
            "sin_imagen_procesada": True,
        })
    return registros


def main():
    if len(sys.argv) < 3:
        print("Uso: python sumar_presapica.py <Base Presapica.xlsx> <catalog_index.json>")
        sys.exit(1)

    excel_path, index_path = sys.argv[1], sys.argv[2]

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    with open(index_path, encoding="utf-8") as f:
        catalogo = json.load(f)
    existentes = {r["sku"] for r in catalogo}

    nuevos = []
    for nombre_hoja, fuente in HOJAS_CON_FOTO.items():
        if nombre_hoja not in wb.sheetnames:
            print(f"Aviso: no encontre la hoja '{nombre_hoja}', la salteo.")
            continue
        regs = procesar_hoja_con_fotos(wb[nombre_hoja], fuente)
        print(f"{nombre_hoja}: {len(regs)} filas procesadas "
              f"({sum(1 for r in regs if not r.get('sin_imagen_procesada'))} con foto)")
        nuevos.extend(regs)

    if "PRESS27" in wb.sheetnames:
        regs = procesar_press27(wb["PRESS27"])
        print(f"PRESS27: {len(regs)} filas procesadas (sin fotos, solo referencia de proveedor/costo)")
        nuevos.extend(regs)

    agregados = 0
    for r in nuevos:
        if r["sku"] in existentes:
            continue
        catalogo.append(r)
        existentes.add(r["sku"])
        agregados += 1

    with open(index_path, "w", encoding="utf-8") as f:
        json.dump(catalogo, f, ensure_ascii=False)

    print(f"\nSe agregaron {agregados} registros nuevos de Base Presapica.")
    print(f"Total en catalog_index.json ahora: {len(catalogo)}")


if __name__ == "__main__":
    main()
