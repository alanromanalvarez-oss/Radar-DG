"""
Radar DG - sumar costo, precio, margen y temporada al catalogo
=================================================================
Cruza catalog_index.json (que ya tiene las fotos procesadas) contra el
Excel "Modelos con inventario mayor a 100_URL_VISTAS.xlsx" por SKU, y le
agrega: costo, precio, margen, temporada. Tambien actualiza
inventario/ventas/sell_through_pct por si esos numeros cambiaron desde
que se genero catalog_index.json.

Si el Excel trae SKUs que todavia NO estan en catalog_index.json (osea,
productos nuevos), el script intenta bajar la foto "Vista 1" desde la URL
para poder compararlos tambien visualmente. Si tu computadora no tiene
internet en ese momento o la URL falla, el SKU se agrega igual (con sus
datos de costo/precio) pero sin foto -- no va a competir en las
comparaciones visuales hasta que se le pueda sumar una foto despues.

Uso:
    python enriquecer_con_costos.py "Modelos con inventario mayor a 100_URL_VISTAS.xlsx" catalog_index.json
"""
import sys
import json
import base64
import io
import re
import openpyxl

try:
    import requests
except ImportError:
    requests = None

try:
    from PIL import Image
    import imagehash
except ImportError:
    Image = None
    imagehash = None

THUMB_SIZE = (220, 220)
JPEG_QUALITY = 70


def normaliza_categoria(cat: str) -> str:
    fix = {
        "BOTÃN": "BOTÍN", "BOTÃ\x8dN": "BOTÍN",
        "MOCASÃN": "MOCASÍN", "MOCASÃ\x8dN": "MOCASÍN",
    }
    return fix.get(cat, cat).title()


def leer_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}

    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx.get("SKU", 0)]:
            continue

        def g(col):
            i = idx.get(col)
            return row[i] if i is not None and i < len(row) else None

        st_val = g("ST")
        st_pct = round(st_val * 100) if isinstance(st_val, (int, float)) and st_val <= 1 else st_val

        vistas = [g(f"Vista {n}") for n in (1, 2, 3, 4)]
        vistas = [v for v in vistas if v]

        filas.append({
            "sku": str(g("SKU")).strip(),
            "categoria": normaliza_categoria(str(g("Categoría") or g("Categoria") or "")),
            "inventario": g("Total INV") or 0,
            "ventas": g("VTA") or 0,
            "sell_through_pct": st_pct or 0,
            "costo": g("Costo"),
            "precio": g("Precio"),
            "margen": g("Margen"),
            "temporada": g("Temporada"),
            "vista_urls": vistas,
        })
    return filas


def descargar_y_hashear(url, timeout=8):
    if not requests or not Image:
        return None
    try:
        r = requests.get(url, timeout=timeout)
        r.raise_for_status()
        im = Image.open(io.BytesIO(r.content)).convert("RGB")
    except Exception as e:
        print(f"    no se pudo bajar {url}: {e}")
        return None

    phash = str(imagehash.phash(im, hash_size=16))
    colorhash = str(imagehash.colorhash(im, binbits=3))
    thumb = im.copy()
    thumb.thumbnail(THUMB_SIZE)
    buf = io.BytesIO()
    thumb.save(buf, format="JPEG", quality=JPEG_QUALITY)
    thumb_b64 = base64.b64encode(buf.getvalue()).decode()
    return phash, colorhash, thumb_b64


def main():
    if len(sys.argv) < 3:
        print("Uso: python enriquecer_con_costos.py <Modelos_URL_VISTAS.xlsx> <catalog_index.json>")
        sys.exit(1)

    excel_path, index_path = sys.argv[1], sys.argv[2]

    with open(index_path, encoding="utf-8") as f:
        catalogo = json.load(f)
    por_sku = {r["sku"]: r for r in catalogo}

    filas = leer_excel(excel_path)
    print(f"Filas en el Excel: {len(filas)}")

    actualizados = 0
    nuevos = 0
    nuevos_sin_foto = 0

    for fila in filas:
        sku = fila["sku"]
        extra = {
            "costo": fila["costo"],
            "precio": fila["precio"],
            "margen": fila["margen"],
            "temporada": fila["temporada"],
            "vista_urls": fila["vista_urls"],
        }
        if sku in por_sku:
            por_sku[sku].update({k: v for k, v in extra.items() if v is not None})
            # refrescamos tambien inventario/ventas/ST por si cambiaron
            por_sku[sku]["inventario"] = fila["inventario"]
            por_sku[sku]["ventas"] = fila["ventas"]
            por_sku[sku]["sell_through_pct"] = fila["sell_through_pct"]
            actualizados += 1
        else:
            nuevo = {
                "sku": sku,
                "categoria": fila["categoria"],
                "inventario": fila["inventario"],
                "ventas": fila["ventas"],
                "sell_through_pct": fila["sell_through_pct"],
                **extra,
            }
            if fila["vista_urls"]:
                resultado = descargar_y_hashear(fila["vista_urls"][0])
                if resultado:
                    nuevo["phash"], nuevo["colorhash"], nuevo["thumb_b64"] = resultado
                else:
                    nuevo["sin_imagen_procesada"] = True
                    nuevos_sin_foto += 1
            por_sku[sku] = nuevo
            nuevos += 1

    catalogo_final = list(por_sku.values())
    with open(index_path, "w", encoding="utf-8") as f:
        json.dump(catalogo_final, f, ensure_ascii=False)

    print(f"\nActualizados (ya existian, se les sumo costo/precio/margen): {actualizados}")
    print(f"Nuevos SKUs agregados: {nuevos} (de los cuales {nuevos_sin_foto} quedaron sin foto procesada)")
    print(f"Total en catalog_index.json ahora: {len(catalogo_final)}")
    if nuevos_sin_foto:
        print("\nOJO: los SKUs sin foto procesada no van a competir en las comparaciones")
        print("visuales todavia. Volve a correr este script desde una compu con buena")
        print("conexion a internet para intentar bajarles la foto.")


if __name__ == "__main__":
    main()
