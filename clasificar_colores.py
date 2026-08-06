"""
Radar DG - clasificar_colores.py
==============================================================================
Asigna a cada SKU una FAMILIA DE COLOR primaria, para el Mapa del catalogo.

De donde sale el color (v28): de los ultimos 3 digitos del SKU, que son el
codigo de color de Dorothy Gaynor (ver "CATALOGO DE COLORES.xlsx"). Se
verifico que el 100% de los 1,395 SKUs estandar del catalogo tienen un codigo
que existe en ese archivo, asi que esta fuente es exacta -- ya no hay que
adivinar el color mirando los pixeles de la foto.

Por que importa: la version anterior estimaba el color desde la imagen y se
equivocaba (ponia cafes en negro y al reves), justo lo que Alan reporto.

Para los SKUs que NO tienen codigo de color (los de Base Presapica: BAIZHEN,
LADY PAU, CAMIDY, etc.) no hay dato, asi que ahi si se estima desde la foto
como antes.

Uso:
    python clasificar_colores.py
"""
import json
import base64
import io
import colorsys
import unicodedata

import numpy as np
from PIL import Image
import openpyxl

CATALOGO = "catalog_index.json"
EXCEL_COLORES = "catalogo_colores.xlsx"
SALIDA = "colores.json"

# Familias primarias. Alan pidio explicitamente NO expandir: un cafe oscuro
# sigue siendo cafe, un negro croco sigue siendo negro.
FAMILIAS = {
    "Negro":      ["NEGRO"],
    "Blanco":     ["BLANCO", "HUESO", "PERLA"],
    "Gris":       ["GRIS", "PLOMO"],
    "Beige":      ["BEIGE", "ARENA", "MAQUILLAJE", "NUDE", "MARFIL", "CREMA"],
    "Café":       ["CAFE", "CAFÉ", "CHOCOLATE", "CAMEL", "LATTE", "COGNAC", "TAN",
                   "TAUPE", "TABACO", "MIEL", "SHEDRON", "PATHE", "MARRON"],
    "Metálico":   ["ORO", "PLATA", "INOX", "METAL", "DORADO", "PLATEADO", "BRONCE"],
    "Rojo":       ["ROJO", "VINO", "BURGUNDY", "GUINDA", "CEREZA"],
    "Rosa":       ["ROSA", "FUCSIA", "FUXIA", "PALO DE ROSA"],
    "Naranja":    ["NARANJA", "CORAL", "TERRACOTA"],
    "Amarillo":   ["AMARILLO", "MOSTAZA", "LIMON"],
    "Verde":      ["VERDE", "OLIVO", "MENTA", "JADE"],
    "Azul":       ["AZUL", "MARINO", "CELESTE", "TURQUESA", "INDIGO"],
    "Morado":     ["MORADO", "LILA", "VIOLETA", "PURPURA"],
    "Multicolor": ["MULTICOLOR", "ANIMAL PRINT", "ESTAMPADO"],
}
# Orden de evaluacion: primero los nombres mas especificos, para que
# "ORO ROSA" caiga en Metalico y no en Rosa, y "ROSA CHAMPAGNE" en Rosa.
PRIORIDAD = ["Multicolor", "Metálico", "Negro", "Blanco", "Gris", "Beige",
             "Café", "Rojo", "Rosa", "Naranja", "Amarillo", "Verde", "Azul", "Morado"]


def normalizar(t):
    t = unicodedata.normalize("NFD", str(t).upper())
    return "".join(c for c in t if unicodedata.category(c) != "Mn").strip()


def familia_desde_nombre(nombre):
    n = normalizar(nombre)
    for fam in PRIORIDAD:
        for clave in FAMILIAS[fam]:
            if normalizar(clave) in n:
                return fam
    return None


# --------- estimacion por pixeles (solo para SKUs sin codigo de color) -------
def _recorte(im, umbral=28, franja=0.03):
    arr = np.array(im.convert("RGB")).astype(np.int16)
    h, w = arr.shape[:2]
    fb = max(2, int(min(h, w) * franja))
    borde = np.concatenate([arr[:fb].reshape(-1, 3), arr[-fb:].reshape(-1, 3),
                            arr[:, :fb].reshape(-1, 3), arr[:, -fb:].reshape(-1, 3)])
    fondo = np.median(borde, axis=0)
    d = np.sqrt(((arr - fondo) ** 2).sum(axis=2))
    ys, xs = np.where(d > umbral)
    if len(xs) == 0:
        return im.convert("RGB"), None
    return im.convert("RGB").crop((xs.min(), ys.min(), xs.max(), ys.max())), fondo


def _fam_hsv(h, s, v):
    if v < 0.20: return "Negro"
    if s < 0.12: return "Blanco" if v > 0.80 else "Gris"
    hd = h * 360
    if s < 0.35 and v > 0.55 and 15 <= hd < 60: return "Beige"
    if v < 0.50 and 10 <= hd < 50: return "Café"
    if hd < 12 or hd >= 340: return "Rosa" if (v > 0.70 and s < 0.50) else "Rojo"
    if 12 <= hd < 40:  return "Café" if v < 0.62 else "Naranja"
    if 40 <= hd < 68:  return "Amarillo"
    if 68 <= hd < 168: return "Verde"
    if 168 <= hd < 258: return "Azul"
    if 258 <= hd < 295: return "Morado"
    return "Rosa"


def color_desde_foto(thumb_b64):
    im = Image.open(io.BytesIO(base64.b64decode(thumb_b64)))
    rec, fondo = _recorte(im)
    rec.thumbnail((90, 90))
    arr = np.array(rec.convert("RGB")).astype(np.float32) / 255.0
    px = arr.reshape(-1, 3)
    if fondo is not None:
        d = np.sqrt(((px - fondo / 255.0) ** 2).sum(axis=1))
        px = px[d > 0.11]
    if len(px) < 30:
        px = arr.reshape(-1, 3)
    cnt = {}
    for r, g, b in px:
        f = _fam_hsv(*colorsys.rgb_to_hsv(r, g, b))
        cnt[f] = cnt.get(f, 0) + 1
    orden = sorted(cnt.items(), key=lambda kv: -kv[1])
    tot = sum(cnt.values())
    if orden[0][1] / tot < 0.38 and len(orden) > 1 and orden[1][1] / tot > 0.30:
        return "Multicolor"
    return orden[0][0]


def main():
    wb = openpyxl.load_workbook(EXCEL_COLORES, read_only=True)
    filas = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))[1:]
    codigos = {str(r[0]).strip().zfill(3): str(r[1]).strip().rstrip(".").strip()
               for r in filas if r[0]}

    catalogo = json.load(open(CATALOGO, encoding="utf-8"))
    salida, por_codigo, por_foto, sin_dato = {}, 0, 0, []
    desconocidos = {}

    for r in catalogo:
        sku = r["sku"]
        fam = None
        if sku.startswith("D") and len(sku) >= 12 and sku[-3:].isdigit():
            nombre = codigos.get(sku[-3:])
            if nombre:
                fam = familia_desde_nombre(nombre)
                if fam:
                    por_codigo += 1
                else:
                    desconocidos[nombre] = desconocidos.get(nombre, 0) + 1
        if fam is None and r.get("thumb_b64"):
            try:
                fam = color_desde_foto(r["thumb_b64"])
                por_foto += 1
            except Exception:
                fam = None
        if fam:
            salida[sku] = fam
        else:
            sin_dato.append(sku)

    json.dump(salida, open(SALIDA, "w", encoding="utf-8"), ensure_ascii=False)

    from collections import Counter
    print(f"Total con color: {len(salida)}")
    print(f"  por codigo de SKU (exacto): {por_codigo}")
    print(f"  estimado desde la foto:     {por_foto}")
    print(f"  sin color:                  {len(sin_dato)}")
    if desconocidos:
        print("\nNombres del Excel que no mapearon a ninguna familia:")
        for k, v in sorted(desconocidos.items(), key=lambda kv: -kv[1]):
            print(f"   {v:4d}  {k}")
    print("\nDistribucion final:")
    for k, v in Counter(salida.values()).most_common():
        print(f"   {v:5d}  {k}")


if __name__ == "__main__":
    main()
