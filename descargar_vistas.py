"""
Radar DG - descargar_vistas.py  (paso 1 de 3 del buscador por embeddings)
==============================================================================
Baja las 4 vistas (angulos) de cada SKU desde el banco de imagenes de Dorothy
Gaynor, leyendo las columnas "Vista 1..4" del Excel
"Modelos con inventario mayor a 100_URL_VISTAS.xlsx".

Por que hace falta: hasta ahora el catalogo tenia UNA sola foto por producto
(la que se saco del PDF). Con 4 angulos por SKU, la foto que saca el comprador
en la feria tiene 4 veces mas chances de parecerse a alguna de las del
catalogo -- ayuda con cualquier tecnica de busqueda, y es la base del indice
de embeddings (ver construir_embeddings.py).

Las fotos se guardan REDIMENSIONADAS (lado maximo 512 px). No hace falta mas
resolucion para comparar siluetas, y asi la carpeta pesa ~200 MB en vez de
varios GB.

Es reanudable: si lo cortas o se cae internet, volve a correr el mismo comando
y saltea las que ya bajo.

Uso:
    python descargar_vistas.py "Modelos con inventario mayor a 100_URL_VISTAS.xlsx" vistas

    (opcional) --sample 20   -> baja solo los primeros 20 SKUs, para probar
"""
import os
import sys
import time
import argparse
import io
import urllib.request
import urllib.error

import openpyxl
from PIL import Image

LADO_MAX = 512
COL_SKU = 0
COLS_VISTA = (15, 16, 17, 18)  # "Vista 1".."Vista 4"
USER_AGENT = "RadarDG/1.0 (indexado interno de catalogo Dorothy Gaynor)"


def descargar_una(url, destino, timeout=30):
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        datos = r.read()
    im = Image.open(io.BytesIO(datos)).convert("RGB")
    im.thumbnail((LADO_MAX, LADO_MAX))
    im.save(destino, format="JPEG", quality=88)


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                  formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("excel", help="ruta al Excel con las columnas Vista 1..4")
    ap.add_argument("carpeta", help="carpeta donde guardar las fotos (se crea sola)")
    ap.add_argument("--sample", type=int, default=None,
                     help="bajar solo los primeros N SKUs (para probar)")
    ap.add_argument("--pausa", type=float, default=0.05,
                     help="segundos de pausa entre descargas (default 0.05)")
    args = ap.parse_args()

    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    os.makedirs(args.carpeta, exist_ok=True)

    wb = openpyxl.load_workbook(args.excel, read_only=True)
    ws = wb[wb.sheetnames[0]]

    filas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        sku = fila[COL_SKU]
        if not sku:
            continue
        urls = [fila[c] for c in COLS_VISTA]
        filas.append((str(sku).strip(), urls))

    if args.sample:
        filas = filas[:args.sample]

    total_urls = sum(1 for _, us in filas for u in us if u and str(u).startswith("http"))
    print(f"SKUs a procesar: {len(filas)} | fotos esperadas: {total_urls}")
    print(f"Guardando en: {os.path.abspath(args.carpeta)}\n")

    bajadas, saltadas, errores = 0, 0, 0
    for i, (sku, urls) in enumerate(filas, 1):
        for n, url in enumerate(urls, 1):
            if not url or not str(url).startswith("http"):
                continue
            destino = os.path.join(args.carpeta, f"{sku}__{n}.jpg")
            if os.path.exists(destino) and os.path.getsize(destino) > 0:
                saltadas += 1
                continue
            try:
                descargar_una(str(url), destino)
                bajadas += 1
                time.sleep(args.pausa)
            except Exception as e:
                errores += 1
                print(f"  [{i}/{len(filas)}] {sku} vista {n}: ERROR -- {e}")

        if i % 50 == 0:
            print(f"  [{i}/{len(filas)}] bajadas: {bajadas} | ya estaban: {saltadas} | errores: {errores}")

    print(f"\nListo. Bajadas ahora: {bajadas} | ya estaban: {saltadas} | errores: {errores}")
    if errores:
        print("Volve a correr el mismo comando para reintentar solo las que fallaron.")


if __name__ == "__main__":
    main()
